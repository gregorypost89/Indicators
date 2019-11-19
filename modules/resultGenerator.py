from openpyxl import load_workbook
from openpyxl.styles import PatternFill as Pf
from openpyxl.formatting.rule import CellIsRule
import pandas as pd

###############################################################################
# Put the paths of the input and output below. Do not replace r or quotes ''. #
inputPath = r'C:\GithubProjects\Indicators\output\test5.xlsx'
outputPath = r'C:\GithubProjects\Indicators\output\xlsxOutput.xlsx'
###############################################################################


darkRedFill = Pf(start_color='FF0000', end_color='FF0000', fill_type='solid')
redFill = Pf(start_color='FF66A4', end_color='FF66A4', fill_type='solid')
pinkFill = Pf(start_color='FFBABA', end_color='FFBABA', fill_type='solid')
ltGreenFill = Pf(start_color='C3FFDF', end_color='C3FFDF', fill_type='solid')
greenFill = Pf(start_color='73FFB4', end_color='73FFB4', fill_type='solid')
darkGreenFill = Pf(start_color='00B956', end_color='00B956', fill_type='solid')

wb = load_workbook(inputPath)
ws = wb.active

maxRow = ws.max_row

gainDict = {}
lossDict = {}
tpMult = 1
slMult = 1.5

for sheet in wb.worksheets:
    sheet.column_dimensions['H'].width = 14
    roc = sheet['G2'].value
    sign = (roc / abs(roc))
    direction = 0
    if sign > 0:
        direction = 'long'
    else:
        direction = 'short'
    tpTarget = round(sheet['C2'].value + (sheet['D2'].value * tpMult * sign), 5)
    slTarget = round(sheet['C2'].value - (sheet['D2'].value * slMult * sign), 5)
    sheet['H2'].value = 'Entering ' + direction + ' at:'
    sheet['H3'].value = str(sheet['C2'].value)
    sheet['H5'].value = 'Take Profit at:'
    sheet['H6'].value = str(tpTarget)
    sheet['H8'].value = 'Stop Loss at:'
    sheet['H9'].value = str(slTarget)
    sheet['I1'].value = 'DistanceToTP'
    sheet['J1'].value = 'DistanceToSL'
    for x in range(2, maxRow):
        iRange, jRange, cRange, bRange = 'I' + str(x), 'J' + str(x), \
                                         'C' + str(x), 'B' + str(x)
        close = sheet[cRange].value
        if close is None:
            continue
        if sign > 0:
            sheet[iRange].value = float(close) - tpTarget
            if sheet[iRange].value > 0:
                result = 'Take Profit'
                if sheet.title in gainDict:
                    continue
                else:
                    gainDict[sheet.title] = [sheet[iRange].value, direction,
                                             sheet[bRange].value]
            sheet[jRange].value = float(close) - slTarget
            if sheet[jRange].value < 0:
                result = 'Stop Loss'
                if sheet.title in lossDict:
                    continue
                else:
                    lossDict[sheet.title] = [sheet[jRange].value, direction,
                                             sheet[bRange].value]

        else:
            sheet[iRange].value = tpTarget - float(close)
            if sheet[iRange].value > 0:
                result = 'Take Profit'
                if sheet.title in gainDict:
                    continue
                else:
                    gainDict[sheet.title] = [sheet[iRange].value, direction,
                                             sheet[bRange].value]
            sheet[jRange].value = slTarget - float(close)
            if sheet[jRange].value < 0:
                result = 'Stop Loss'
                if sheet.title in lossDict:
                    continue
                else:
                    lossDict[sheet.title] = [sheet[jRange].value, direction,
                                             sheet[bRange].value]
    iMax, jMax = 'I' + str(maxRow), 'J' + str(maxRow)
    sheet.conditional_formatting.add('I2:' + str(iMax),
                                     CellIsRule(operator='greaterThan',
                                                formula=[('0')],
                                                stopIfTrue=True,
                                                fill=darkGreenFill))
    sheet.conditional_formatting.add('I2:' + str(iMax),
                                     CellIsRule(operator='between',
                                                formula=[('(0)'),
                                                         ('(H$3 - H$6) * .5')],
                                                stopIfTrue=True,
                                                fill=greenFill))
    sheet.conditional_formatting.add('I2:' + str(iMax),
                                     CellIsRule(operator='between',
                                                formula=[('(H$3 - H$6) * .5'),
                                                         ('H$3 - H$6')],
                                                stopIfTrue=True,
                                                fill=ltGreenFill))
    sheet.conditional_formatting.add('J2:' + str(jMax),
                                     CellIsRule(operator='lessThan',
                                                formula=['0'],
                                                stopIfTrue=True,
                                                fill=darkRedFill))
    sheet.conditional_formatting.add('J2:' + str(jMax),
                                     CellIsRule(operator='between',
                                                formula=['0',
                                                         ('(H$3 - H$9) * .5')],
                                                stopIfTrue=True,
                                                fill=redFill))
    sheet.conditional_formatting.add('J2:' + str(jMax),
                                     CellIsRule(operator='between',
                                                formula=[('(H$3 - H$9) * .5'),
                                                         ('(H$3 - H$9) * 1')],
                                                stopIfTrue=True,
                                                fill=pinkFill))

wb.create_sheet("summary", 0)
ws = wb.active
ws.sheet_properties.tabColor = "1072BA"
wb.save(outputPath)

gainDF = pd.DataFrame.from_dict(gainDict, orient='index')
lossDF = pd.DataFrame.from_dict(lossDict, orient='index')
mergedDF = gainDF.join(lossDF, lsuffix='a', rsuffix='b').rename(
    columns={'0a': 'tpGain', '1a': 'tpDirection', '2a': 'tpDate',
             '0b': 'slLoss', '1b': 'slDirection', '2b': 'slDate'})

mergedDF.to_excel(r'C:\GithubProjects\Indicators\output\summaryOutput.xlsx')
wb2 = load_workbook(r'C:\GithubProjects\Indicators\output\summaryOutput.xlsx')
ws2 = wb2.active
ws2.column_dimensions['I'].width, ws2.column_dimensions['K'].width = 14, 14
maxRow2 = ws2.max_row
profits, losses, total = 0, 0, 0
for y in range(2, maxRow2+1):
    bRange, cRange, dRange, eRange, fRange, gRange = 'B' + str(y), 'C' + str(y), \
                                                     'D' + str(y), 'E' + str(y), \
                                                     'F' + str(y), 'G' + str(y)
    if ws2[dRange].value is None:
        ws2[dRange].value = 0
    if ws2[gRange].value is None:
        ws2[gRange].value = 0
    tpDate, slDate = int(str(ws2[dRange].value).replace('.', '')), \
                     int(str(ws2[gRange].value).replace('.', ''))
    if tpDate < slDate or slDate is 0:
        ws2[bRange].fill = Pf("solid", fgColor="2f7a30") #green
        profits += 1
    if tpDate > slDate and slDate is not 0:
        ws2[eRange].fill = Pf("solid", fgColor="a8324a") #red
        losses += 1
    total += 1

print(profits)
print(losses)
print(total)
ws2['I1'].value = 'Take Profit'
ws2['I2'].value = 'Total Hits'
ws2['I3'].value = 'Total Entries'
ws2['I4'].value = 'Profit Rate'
ws2['J2'].value = profits
ws2['J3'].value = total
ws2['J4'].value = (profits/total) * 100
ws2['K1'].value = 'Stop Loss'
ws2['K2'].value = 'Total Hits'
ws2['K3'].value = 'Total Entries'
ws2['K4'].value = 'Loss Rate'
ws2['L2'].value = losses
ws2['L3'].value = total
ws2['L4'].value = (losses/total) * 100
wb2.save(r'C:\GithubProjects\Indicators\output\summaryOutput2.xlsx')

    # sheet.conditional_formatting.add('B' + str(y)), CellIsRule(
    #                                             operator='greaterThan',
    #                                             formula=[('0')],
    #                                             stopIfTrue=True,
    #                                             fill=darkGreenFill))

#ws2.column_dimensions.width = 14

#my input: C:\GithubProjects\Indicators\output\test5.xlsx
#my output: C:\GithubProjects\Indicators\output\xlsxOutput.xlsx'
